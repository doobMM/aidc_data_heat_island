# -*- coding: utf-8 -*-
"""정전(지사) → 시군구 매핑 + 정전 × 온열질환 × 산업전력 3중 교차 (2026-06-22)
입력(정보공개청구/): 0622 정전 xlsx, derived 온열질환 csv, (0621) 산업전력 xlsx 4개
출력: 콘솔 결과 + 지사_시군구_crosswalk csv
"""
import openpyxl, pandas as pd, glob, re, os
from collections import defaultdict
BASE = r"C:/cross_the_street/docs/research/_data-center/기획서_꾸러미/정보공개청구"

def collapse(k):
    # "경기도 안산시 단원구" -> "경기도 안산시" (도-소속 시의 구만 통합; 광역시 구는 유지)
    return re.sub(r'^(\S*도\s+\S+시)\s+\S+구$', r'\1', str(k).strip())

# 1) 정전: (본부,지사) 집계
df = pd.read_excel(os.path.join(BASE, "FOIA_0622_한전_정전이력_2020-2025.xlsx"), sheet_name="최종")
df['발생'] = pd.to_datetime(df['발생일시'], errors='coerce')
df['복구'] = pd.to_datetime(df['복구일시'], errors='coerce')
df['dur_h'] = (df['복구'] - df['발생']).dt.total_seconds()/3600
df['호수'] = pd.to_numeric(df['정전호수'], errors='coerce')
gj = df.groupby(['본부','지사']).agg(정전건수=('순번','size'),
        일시건수=('정전구분', lambda s:(s=='일시').sum()),
        정전호수=('호수','sum'), 평균지속h=('dur_h','mean')).reset_index()

# 2) 온열질환 + 전력 (같은 "시도 시군구" 키, 구→시 통합)
heat = pd.read_csv(os.path.join(BASE, "DERIVED_0614_시군구_온열질환_집계.csv"))
heat['key'] = heat['시군구'].map(collapse)
heatg = heat.groupby('key').agg(온열발생=('발생','sum'), 중증=('중증','sum')).reset_index()
heatg['중증률'] = (heatg['중증']/heatg['온열발생']*100).round(1)
KEYS = set(heatg['key'])

elec = defaultdict(float)
for f in sorted(glob.glob(os.path.join(BASE, "OPEN_0621_한전_산업분류법정동전력_*.xlsx"))):
    wb = openpyxl.load_workbook(f, read_only=True); ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True); next(it)
    for r in it:
        if r[10] is None: continue
        elec[collapse(f"{r[2]} {r[3]}")] += float(r[10])
    wb.close()

# 3) 지사 -> key 매핑 (base 부분일치 + 본부 지역 disambiguation)
SIDO = sorted({k.split()[0] for k in KEYS})
def bonbu_sido(b):
    out=[]
    for sd in SIDO:
        short = sd[:2]
        if short in str(b): out.append(sd)
    return out
def bare(k): return k.split()[-1]
# v2: 공개 관할구역 기반 수동 매핑 — 도-소속 시의 방위분할·이름변경만 (광역시 복수구·본부직할은 구조상 제외)
OVERRIDE = {
    '동용인지사':'경기도 용인시','서용인지사':'경기도 용인시',
    '서수원지사':'경기도 수원시','서평택지사':'경기도 평택시',
    '동청주지사':'충청북도 청주시','남전주지사':'전북특별자치도 전주시',
    '북포항지사':'경상북도 포항시','강릉특별지사':'강원특별자치도 강릉시',
    '마산지사':'경상남도 창원시','진해지사':'경상남도 창원시',
}
def map_jisa(bonbu, jisa):
    if jisa in OVERRIDE: return OVERRIDE[jisa] if OVERRIDE[jisa] in KEYS else None
    base = re.sub(r'(지사|전력)$','', str(jisa)).strip()
    if not base: return None
    cands = [k for k in KEYS if base in bare(k)]
    if len(cands)==1: return cands[0]
    if len(cands)>1:
        reg = bonbu_sido(bonbu)
        f = [k for k in cands if k.split()[0] in reg]
        if len(f)>=1: return f[0]
        return cands[0]
    return None

gj['key'] = [map_jisa(b,j) for b,j in zip(gj['본부'], gj['지사'])]
mapped = gj.dropna(subset=['key'])
rate = len(mapped)/len(gj)*100
print(f"지사 총 {len(gj)}개 / 매핑 {len(mapped)}개 ({rate:.1f}%)")
print(f"매핑 실패 {int(gj['key'].isna().sum())}개:", list(gj[gj['key'].isna()]['지사']))

# key 단위 정전 합산
elecdf = pd.DataFrame([(k,v) for k,v in elec.items()], columns=['key','전력GWh'])
elecdf['전력GWh'] = elecdf['전력GWh']/1e6
outk = mapped.groupby('key').agg(정전건수=('정전건수','sum'), 정전호수=('정전호수','sum'),
        평균지속h=('평균지속h','mean')).reset_index()

m = outk.merge(heatg[['key','온열발생','중증률']], on='key', how='inner').merge(elecdf, on='key', how='left')
print(f"\n3중 교차 매칭 시군구: {len(m)}개")
def corr(a,b): return m[a].corr(m[b])
print(f"정전건수 vs 온열발생  r = {corr('정전건수','온열발생'):.2f}")
print(f"정전호수 vs 온열발생  r = {corr('정전호수','온열발생'):.2f}")
print(f"정전건수 vs 전력GWh   r = {corr('정전건수','전력GWh'):.2f}")
print(f"온열발생 vs 전력GWh   r = {corr('온열발생','전력GWh'):.2f}")

print("\n[정전건수 Top 15 시군구 — 온열·전력 동반]")
top = m.sort_values('정전건수', ascending=False).head(15)
for _,x in top.iterrows():
    print(f"  정전{int(x.정전건수):>4}건 호수{int(x.정전호수 or 0):>7} 지속{x.평균지속h:>4.1f}h | 온열{int(x.온열발생):>4} 중증{x.중증률:>4.1f}% | 전력{(x.전력GWh or 0):>7,.0f}GWh  {x.key}")

print("\n[온열발생 Top 15 시군구 — 정전 동반]")
for _,x in m.sort_values('온열발생', ascending=False).head(15).iterrows():
    print(f"  온열{int(x.온열발생):>4} | 정전{int(x.정전건수):>4}건 지속{x.평균지속h:>4.1f}h | 전력{(x.전력GWh or 0):>7,.0f}GWh  {x.key}")

# crosswalk 저장
gj.to_csv(os.path.join(BASE,"DERIVED_0630_지사_시군구_crosswalk_v2.csv"), index=False, encoding='utf-8-sig')
m.to_csv(os.path.join(BASE,"DERIVED_0630_정전_시군구_3중_merged_v2.csv"), index=False, encoding='utf-8-sig')
print("\n저장: 지사_시군구_crosswalk / 정전_시군구_3중_merged")
print("DONE-CROSS")
